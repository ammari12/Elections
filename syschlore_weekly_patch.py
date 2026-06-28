"""
══════════════════════════════════════════════════════════════════════════════
SysChlore v2.0 — Weekly Patch Module
Mise à jour hebdomadaire automatique de la base de données accidents chlore
══════════════════════════════════════════════════════════════════════════════

USAGE:
    python syschlore_weekly_patch.py                    # mode interactif
    python syschlore_weekly_patch.py --auto             # mode silencieux (cron)
    python syschlore_weekly_patch.py --dry-run          # simulation sans écriture

CRON (exécution chaque lundi à 06h00):
    0 6 * * 1 cd /chemin/syschlore && python syschlore_weekly_patch.py --auto

DÉPENDANCES:
    pip install requests beautifulsoup4 anthropic openpyxl pandas python-dotenv
"""

import os
import sys
import json
import math
import logging
import hashlib
import argparse
import textwrap
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup

# ── Anthropic Claude pour extraction structurée ───────────────────────────────
try:
    import anthropic
    CLAUDE_AVAILABLE = True
except ImportError:
    CLAUDE_AVAILABLE = False
    print("⚠  Module anthropic non installé. Extraction automatique désactivée.")
    print("   pip install anthropic")

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("syschlore_patch.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("syschlore_patch")

# ══════════════════════════════════════════════════════════════════════════════
# 1. CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

BASE_DIR   = Path(__file__).parent
DATA_PATH  = BASE_DIR / "data" / "accidents_chlore_rectifie.xlsx"
CACHE_PATH = BASE_DIR / "data" / "patch_cache.json"
LOG_PATH   = BASE_DIR / "data" / "patch_history.json"

ERPG = {"ERPG-1": 1.0, "ERPG-2": 3.0, "ERPG-3": 20.0}
PG_COEFF = {
    "A": dict(ay=0.22, by=0.89, az=0.20, bz=0.89),
    "B": dict(ay=0.16, by=0.86, az=0.12, bz=0.86),
    "C": dict(ay=0.11, by=0.83, az=0.08, bz=0.82),
    "D": dict(ay=0.08, by=0.85, az=0.06, bz=0.80),
    "E": dict(ay=0.06, by=0.81, az=0.03, bz=0.76),
    "F": dict(ay=0.04, by=0.75, az=0.016, bz=0.72),
}
STAB_MAP = {"A":"A","B":"B","C":"C","D":"D","E":"E","F":"F",
            "Stable":"D","Neutre":"D","Instable":"A"}

F_SITE = {
    "Zone rurale": 0.05, "Route (zone rurale)": 0.05, "Voies ferrées rurales": 0.06,
    "Station eau": 0.06, "Station eau rurale": 0.06, "Station eau urbaine": 0.07,
    "Station potabilisation": 0.07, "Réseau eau": 0.07, "Canalisation": 0.07,
    "Usine chimique": 0.08, "Usine chimique / pompe": 0.08, "Usine papier": 0.09,
    "Évaporateur chimique": 0.09, "Zone industrielle": 0.09, "Réservoir industriel": 0.10,
    "Complexe BASF": 0.10, "Entreposage cylindres": 0.10,
    "Port semi-ouvert": 0.12, "Port": 0.12, "Port urbain": 0.14,
    "Port industriel": 0.13, "Port / chargement": 0.13, "Port / bateau": 0.12,
    "Voie ferrée (zone mixte)": 0.25, "Voies ferrées": 0.25, "Route": 0.20,
    "Échangeur": 0.22, "Zone urbaine": 0.20, "Réservoir urbain": 0.18,
    "Zone de recyclage": 0.15, "Hôpital": 0.25,
    "Piscine municipale": 0.35, "Piscine": 0.35, "Piscine hôtel": 0.35,
    "Piscine / local technique": 0.35, "Piscine / remplissage": 0.35,
    "Centre aquatique": 0.35,
}

# Sources de surveillance
SOURCES = [
    {
        "name": "ARIA / BARPI",
        "url": "https://www.aria.developpement-durable.gouv.fr/recherche/?q=chlore",
        "lang": "fr",
    },
    {
        "name": "ARIA Recent",
        "url": "https://www.aria.developpement-durable.gouv.fr/accidents-recents/",
        "lang": "fr",
    },
    {
        "name": "CSB (USA)",
        "url": "https://www.csb.gov/investigations/",
        "lang": "en",
    },
    {
        "name": "OSHA Incidents",
        "url": "https://www.osha.gov/pls/imis/accidentsearch.html",
        "lang": "en",
    },
    {
        "name": "ReliefWeb Hazmat",
        "url": "https://reliefweb.int/disasters?search=chlorine",
        "lang": "en",
    },
    {
        "name": "Google News Chlorine",
        "url": "https://news.google.com/rss/search?q=chlorine+accident+leak+spill&hl=en",
        "lang": "en",
        "rss": True,
    },
    {
        "name": "Google News Chlore FR",
        "url": "https://news.google.com/rss/search?q=accident+chlore+fuite&hl=fr",
        "lang": "fr",
        "rss": True,
    },
]


# ══════════════════════════════════════════════════════════════════════════════
# 2. MOTEUR PHYSIQUE SysChlore (copie autonome)
# ══════════════════════════════════════════════════════════════════════════════

def sigma_pg(x: float, stab: str) -> tuple[float, float]:
    """Coefficients de dispersion Pasquill-Gifford (Turner, 1994)."""
    c = PG_COEFF.get(stab, PG_COEFF["D"])
    sy = c["ay"] * x ** c["by"]
    sz = max(c["az"] * x ** c["bz"], 0.1)
    return sy, sz


def h_effectif(H: float, type_lib: str) -> float:
    """Hauteur effective : 0 pour les libérations brutales (gaz dense)."""
    return 0.0 if type_lib == "Brutale" else H


def q_debit_kgs(Q_kg: float, duree_min: float, type_lib: str) -> float:
    """Débit massique effectif en kg/s."""
    if type_lib == "Brutale":
        return Q_kg / max(duree_min, 1) / 60
    return Q_kg / max(duree_min, 1) / 60


def conc_axiale_ppm(Q_kgs: float, u: float, x: float, stab: str, H_eff: float = 0.0) -> float:
    """Concentration axiale (ppm) par modèle gaussien PG."""
    sy, sz = sigma_pg(x, stab)
    denom = math.pi * u * sy * sz
    if denom <= 0:
        return 0.0
    C_mg = (Q_kgs * 1000) / denom * math.exp(-0.5 * (H_eff / sz) ** 2 if H_eff > 0 else 0)
    return C_mg * (24.45 / 70.9)


def rayon_seuil(Q_kg: float, u: float, stab: str, H: float,
                seuil_ppm: float, duree_min: float, type_lib: str) -> float:
    """Rayon de danger (m) pour un seuil ERPG donné, avec correction puff."""
    Q_kgs  = q_debit_kgs(Q_kg, duree_min, type_lib)
    H_eff  = h_effectif(H, type_lib)
    R = 0.0
    for x in range(10, 6010, 10):
        if conc_axiale_ppm(Q_kgs, u, x, stab, H_eff) < seuil_ppm:
            R = x
            break
    # Correction puff pour libérations courtes
    if type_lib == "Brutale" and R > 0:
        ts = duree_min * 60
        R_plume = R
        if ts < R_plume / u:
            R = math.sqrt(R_plume * u * ts)
    return R


def simulation_monte_carlo(Q_kg: float, u: float, stab: str, H: float,
                            duree_min: float, type_lib: str,
                            dist_pop: float, N: int = 2000) -> dict:
    """
    Simulation Monte Carlo (N itérations).
    Retourne : P(C>ERPG-k) à la distance de la population, niveau d'alerte,
               médiane et IC 90 %.
    """
    np.random.seed(42)
    Q_kgs  = q_debit_kgs(Q_kg, duree_min, type_lib)
    H_eff  = h_effectif(H, type_lib)
    Qv = np.random.lognormal(np.log(max(Q_kgs, 1e-6)), 0.30, N)
    uv = np.clip(np.random.lognormal(np.log(max(u, 0.3)), 0.20, N), 0.3, 15.0)
    sy, sz = sigma_pg(dist_pop, stab)
    denom_v = math.pi * uv * sy * sz
    C_v = np.where(denom_v > 0, (Qv * 1000) / denom_v * (24.45 / 70.9), 0.0)
    if H_eff > 0:
        C_v *= math.exp(-0.5 * (H_eff / sz) ** 2)
    C_v = np.clip(C_v, 0, None)
    p_e1 = float((C_v > ERPG["ERPG-1"]).mean())
    p_e2 = float((C_v > ERPG["ERPG-2"]).mean())
    p_e3 = float((C_v > ERPG["ERPG-3"]).mean())
    # Niveau d'alerte MC
    if p_e2 >= 0.70 or p_e3 >= 0.30:
        niveau = "ROUGE"
    elif p_e2 >= 0.30:
        niveau = "ORANGE"
    elif p_e2 >= 0.10:
        niveau = "JAUNE"
    else:
        niveau = "VERT"
    # Déterministe
    C_det = conc_axiale_ppm(Q_kgs, u, dist_pop, stab, H_eff)
    if C_det >= ERPG["ERPG-3"]:
        niveau_det = "ROUGE"
    elif C_det >= ERPG["ERPG-2"]:
        niveau_det = "ORANGE"
    elif C_det >= ERPG["ERPG-1"]:
        niveau_det = "JAUNE"
    else:
        niveau_det = "VERT"
    return {
        "C_det_ppm": round(C_det, 3),
        "niveau_det": niveau_det,
        "p_e1": round(p_e1, 3),
        "p_e2": round(p_e2, 3),
        "p_e3": round(p_e3, 3),
        "niveau_mc": niveau,
        "p50": round(float(np.median(C_v)), 2),
        "p05": round(float(np.percentile(C_v, 5)), 2),
        "p95": round(float(np.percentile(C_v, 95)), 2),
        "paradoxe_detecte": (niveau_det in ["VERT", "JAUNE"] and niveau in ["ORANGE", "ROUGE"]),
    }


def calcul_indice_G(Q_kg: float, u: float, stab: str, H: float,
                    duree_min: float, type_lib: str, dist_pop: float,
                    dens_pop: float, config_site: str,
                    alerte: str, epi_num: int, delai_evac: float,
                    coord: int, capa_med_num: int) -> float:
    """Indice de gravité composite G ∈ [0, 10] — SysChlore v2.0."""
    Q_kgs  = q_debit_kgs(Q_kg, duree_min, type_lib)
    H_eff  = h_effectif(H, type_lib)
    C_pop  = min(conc_axiale_ppm(Q_kgs, u, max(dist_pop, 10), stab, H_eff), 500.0)
    f_site = F_SITE.get(config_site, 0.12)
    # Scores normalisés
    expo   = max(math.log10(C_pop + 0.01) / math.log10(500.01), 0.0)
    q_sc   = max(0.0, min(1.0, (math.log10(max(Q_kg, 1)) - 1) / (math.log10(60000) - 1)))
    dur_sc = min(duree_min / 120.0, 1.0)
    raw    = expo * 5.5 + q_sc * 2.5 + dur_sc * 1.0
    # Multiplicateurs contextuels
    brutal_m  = 1.30 if type_lib == "Brutale" else 1.00
    stab_m    = {"A": 0.70, "B": 0.80, "C": 0.90, "D": 1.00, "E": 1.10, "F": 1.20}.get(
                  STAB_MAP.get(stab, "D"), 1.00)
    dens_m    = 0.80 + min(dens_pop / 10000, 0.60)
    site_m    = 0.70 + 1.20 * f_site
    alerte_m  = 0.78 if alerte == "Oui" else 1.00
    epi_m     = {1: 1.20, 2: 1.05, 3: 0.90, 4: 0.75}.get(epi_num, 1.00)
    evac_m    = min(1.0 + delai_evac / 120.0, 1.40)
    coord_m   = {1: 1.15, 2: 1.05, 3: 0.95, 4: 0.85}.get(coord, 1.00)
    capa_m    = {1: 1.10, 2: 1.00, 3: 0.85}.get(capa_med_num, 1.00)
    score = raw * brutal_m * stab_m * dens_m * site_m
    score *= alerte_m * epi_m * evac_m * coord_m * capa_m
    return round(min(10.0, max(0.0, score)), 2)


# ══════════════════════════════════════════════════════════════════════════════
# 3. SCRAPING WEB
# ══════════════════════════════════════════════════════════════════════════════

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
}


def fetch_page(url: str, timeout: int = 15) -> str | None:
    """Récupère le texte HTML d'une URL. Retourne None si erreur."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r.text
    except Exception as e:
        log.warning(f"Fetch failed {url}: {e}")
        return None


def parse_rss(html: str) -> list[dict]:
    """Parse un flux RSS et extrait titre + lien + date."""
    soup = BeautifulSoup(html, "xml")
    items = []
    for item in soup.find_all("item")[:20]:
        title = item.find("title")
        link  = item.find("link")
        pub   = item.find("pubDate")
        if title and link:
            items.append({
                "title": title.get_text(strip=True),
                "url": link.get_text(strip=True),
                "date": pub.get_text(strip=True) if pub else "",
                "snippet": "",
            })
    return items


def parse_html_articles(html: str, url: str) -> list[dict]:
    """Extrait les articles potentiellement pertinents depuis un HTML générique."""
    soup = BeautifulSoup(html, "html.parser")
    keywords = ["chlore", "chlorine", "cl2", "cl₂", "leak", "spill",
                 "fuite", "accident", "incident", "toxic", "hazmat"]
    articles = []
    for tag in soup.find_all(["article", "div", "li", "tr"],
                              class_=lambda c: c and any(
                                  k in c.lower() for k in
                                  ["article", "result", "item", "post", "event", "accident"]
                              )):
        text = tag.get_text(" ", strip=True).lower()
        if not any(k in text for k in keywords):
            continue
        a_tag = tag.find("a", href=True)
        if not a_tag:
            continue
        href = a_tag.get("href", "")
        if href.startswith("/"):
            from urllib.parse import urlparse
            parsed = urlparse(url)
            href = f"{parsed.scheme}://{parsed.netloc}{href}"
        articles.append({
            "title": a_tag.get_text(strip=True)[:200],
            "url": href,
            "date": "",
            "snippet": text[:400],
        })
    return articles[:20]


def scrape_all_sources() -> list[dict]:
    """Scrape toutes les sources et retourne la liste brute des candidats."""
    candidates = []
    for src in SOURCES:
        log.info(f"Scraping: {src['name']}")
        html = fetch_page(src["url"])
        if not html:
            continue
        if src.get("rss"):
            items = parse_rss(html)
        else:
            items = parse_html_articles(html, src["url"])
        for it in items:
            it["source"] = src["name"]
        candidates.extend(items)
        log.info(f"  → {len(items)} candidats trouvés")
    log.info(f"Total candidats bruts: {len(candidates)}")
    return candidates


def filter_chlorine_accidents(candidates: list[dict]) -> list[dict]:
    """Filtre sur les mots-clés chlore et élimine les doublons."""
    keywords_pos = ["chlore", "chlorine", "cl2", "cl₂", "hypochlorite",
                    "fuite", "leak", "spill", "rejet", "intoxication",
                    "accident chimique", "chemical accident", "hazmat chlor"]
    keywords_neg = ["swimming pool chlorine tablet", "chlorinated water treatment",
                    "pool chemical", "household", "piscine entretien", "nettoyage"]
    cache = load_cache()
    seen_hashes = set(cache.get("seen", []))
    filtered = []
    for c in candidates:
        text = (c.get("title", "") + " " + c.get("snippet", "")).lower()
        if not any(k in text for k in keywords_pos):
            continue
        if any(k in text for k in keywords_neg):
            continue
        # Déduplication par hash URL + titre
        h = hashlib.md5((c.get("url", "") + c.get("title", "")).encode()).hexdigest()
        if h in seen_hashes:
            continue
        c["_hash"] = h
        filtered.append(c)
    log.info(f"Après filtrage: {len(filtered)} accidents potentiels nouveaux")
    return filtered


# ══════════════════════════════════════════════════════════════════════════════
# 4. EXTRACTION STRUCTURÉE VIA CLAUDE
# ══════════════════════════════════════════════════════════════════════════════

EXTRACTION_PROMPT = """
Tu es un expert en analyse d'accidents industriels au chlore (Cl₂).
On te donne le texte d'un article de presse ou d'un rapport sur un incident.
Extrais les informations dans un JSON STRICT avec EXACTEMENT ces champs.
Si une valeur est inconnue, mets null (jamais une chaîne vide).

JSON à retourner (et rien d'autre, pas de markdown, pas de backticks) :
{
  "Nom_Accident": "Lieu Année",
  "Année": 2024,
  "Pays": "Nom du pays",
  "Quantité_Cl2_kg": null_ou_nombre,
  "Type_libération": "Brutale" ou "Progressive",
  "Hauteur_source_m": null_ou_nombre,
  "Durée_libération_min": null_ou_nombre,
  "Vitesse_vent_ms": null_ou_nombre,
  "Stabilité_atm": "Neutre" ou "Stable",
  "Configuration_site": "choisir parmi: Zone rurale / Route / Usine chimique / Zone industrielle / Zone urbaine / Station eau / Port / Port semi-ouvert / Voie ferrée / Piscine / Hôpital / Réservoir industriel",
  "Obstacles": "Oui" ou "Non",
  "Densité_pop_km2": null_ou_nombre,
  "Distance_pop_m": null_ou_nombre,
  "Durée_expo_min": null_ou_nombre,
  "Niveau_EPI": "Très faible" ou "Faible" ou "Moyen",
  "Alerte_précoce": "Oui" ou "Non",
  "Délai_évacuation_min": null_ou_nombre,
  "Formation_pers": null_ou_entier_1_à_4,
  "Coord_secours": null_ou_entier_1_à_4,
  "Délai_intervention_min": null_ou_nombre,
  "Capacité_médicale": "Limitée" ou "Moyenne" ou "Bonne",
  "Décès": null_ou_entier,
  "Blessés_graves": null_ou_entier,
  "Blessés_légers": null_ou_entier,
  "Blessés_total": null_ou_entier,
  "Zone_impact_km2": null_ou_nombre,
  "Périmètre_évacuation_m": null_ou_nombre,
  "Impact_environnement": "Faible" ou "Modéré" ou "Élevé",
  "Contamination_eau": "Oui" ou "Non",
  "Contamination_sol": "Oui" ou "Non",
  "Coût_MUSD": null_ou_nombre,
  "Jours_arrêt": null_ou_nombre,
  "source_url": "url de l'article",
  "confiance": "Haute" ou "Moyenne" ou "Faible"
}

RÈGLES D'INFÉRENCE (quand non précisé) :
- Durée_libération_min : 2 si Brutale, 60 si Progressive
- Vitesse_vent_ms : 2.0 si non précisé
- Stabilité_atm : "Neutre" si de jour, "Stable" si la nuit
- Hauteur_source_m : 0 si Brutale, 5 si Progressive
- Formation_pers : 1=aucune, 2=basique, 3=correcte, 4=excellente
- Coord_secours : 1=absente, 2=faible, 3=correcte, 4=excellente
- Capacité_médicale : Limitée si < 50 000 hab, Bonne si > 500 000 hab
- Obstacles : "Oui" si zone urbaine ou portuaire, "Non" si rural

ARTICLE :
{article_text}
"""


def fetch_article_text(url: str) -> str:
    """Récupère le texte principal d'un article (heuristique)."""
    html = fetch_page(url, timeout=20)
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    # Retirer scripts, styles, nav
    for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
        tag.decompose()
    # Trouver le bloc principal
    for selector in ["article", "main", ".content", ".article-body",
                     ".post-content", "#content", ".entry-content"]:
        block = soup.select_one(selector)
        if block:
            return block.get_text(" ", strip=True)[:4000]
    return soup.get_text(" ", strip=True)[:4000]


def extract_with_claude(article: dict) -> dict | None:
    """Utilise Claude pour extraire les champs structurés d'un article."""
    if not CLAUDE_AVAILABLE:
        return None
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        log.warning("ANTHROPIC_API_KEY non définie — extraction Claude désactivée")
        return None
    text = fetch_article_text(article.get("url", ""))
    if not text:
        text = article.get("title", "") + " " + article.get("snippet", "")
    if len(text) < 50:
        log.warning(f"Texte trop court pour extraction: {article.get('url','')}")
        return None
    prompt = EXTRACTION_PROMPT.format(article_text=text[:3500])
    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1200,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        # Nettoyer backticks si présents
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        data = json.loads(raw)
        data["source_url"] = article.get("url", "")
        data["_hash"] = article.get("_hash", "")
        log.info(f"  ✓ Extraction Claude réussie: {data.get('Nom_Accident','?')}")
        return data
    except Exception as e:
        log.error(f"Extraction Claude échouée pour {article.get('url','')}: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
# 5. CALCUL DES CHAMPS DÉRIVÉS (OUTPUTS SysChlore)
# ══════════════════════════════════════════════════════════════════════════════

ENUM_MAPS = {
    "Type_libération_num":     {"Brutale": 1, "Progressive": 0},
    "Stabilité_atm_code":      {"A":1,"B":2,"C":3,"D":4,"E":5,"F":6,"Neutre":4,"Stable":5},
    "Obstacles_bin":           {"Oui": 1, "Non": 0},
    "Niveau_EPI_num":          {"Très faible": 1, "Faible": 2, "Moyen": 3},
    "Alerte_précoce_bin":      {"Oui": 1, "Non": 0},
    "Capacité_médicale_num":   {"Limitée": 1, "Moyenne": 2, "Bonne": 3},
    "Impact_environnement_num":{"Faible": 1, "Modéré": 2, "Élevé": 3},
    "Contamination_eau_bin":   {"Oui": 1, "Non": 0},
    "Contamination_sol_bin":   {"Oui": 1, "Non": 0},
}


def completer_champs_derives(d: dict) -> dict:
    """
    Calcule tous les champs numériques binaires/ordinaux et les outputs SysChlore.
    Remplit les valeurs manquantes par des valeurs par défaut raisonnables.
    """
    # ── Valeurs par défaut si null ──────────────────────────────────────────
    defaults = {
        "Hauteur_source_m":     0,
        "Durée_libération_min": 2 if d.get("Type_libération") == "Brutale" else 60,
        "Vitesse_vent_ms":      2.0,
        "Stabilité_atm":        "Neutre",
        "Configuration_site":   "Zone industrielle",
        "Obstacles":            "Non",
        "Densité_pop_km2":      1000,
        "Distance_pop_m":       200,
        "Durée_expo_min":       30,
        "Niveau_EPI":           "Faible",
        "Alerte_précoce":       "Non",
        "Délai_évacuation_min": 30,
        "Formation_pers":       2,
        "Coord_secours":        2,
        "Délai_intervention_min": 20,
        "Capacité_médicale":    "Moyenne",
        "Blessés_graves":       0,
        "Blessés_légers":       0,
        "Impact_environnement": "Modéré",
        "Contamination_eau":    "Non",
        "Contamination_sol":    "Non",
        "Coût_MUSD":            None,
        "Jours_arrêt":          None,
    }
    for k, v in defaults.items():
        if d.get(k) is None:
            d[k] = v

    # ── Champs binaires/numériques ──────────────────────────────────────────
    for col, mapping in ENUM_MAPS.items():
        src_col = col.replace("_num", "").replace("_bin", "").replace("_code", "")
        # Correspondance approximative au nom de colonne source
        for key, val in mapping.items():
            if d.get(src_col) == key or d.get(col.rstrip("_num").rstrip("_bin")) == key:
                d[col] = val
                break
        if col not in d:
            d[col] = 0

    # Blessés total
    if d.get("Blessés_total") is None:
        d["Blessés_total"] = (d.get("Blessés_graves") or 0) + (d.get("Blessés_légers") or 0)

    # ── Paramètres physiques ────────────────────────────────────────────────
    Q_kg      = float(d.get("Quantité_Cl2_kg") or 500)
    u         = float(d.get("Vitesse_vent_ms") or 2.0)
    stab      = STAB_MAP.get(d.get("Stabilité_atm", "Neutre"), "D")
    H         = float(d.get("Hauteur_source_m") or 0)
    dur       = float(d.get("Durée_libération_min") or 60)
    type_lib  = d.get("Type_libération", "Progressive")
    dist_pop  = float(d.get("Distance_pop_m") or 200)
    dens_pop  = float(d.get("Densité_pop_km2") or 1000)
    config    = d.get("Configuration_site", "Zone industrielle")
    alerte    = d.get("Alerte_précoce", "Non")
    epi_num   = int(d.get("Niveau_EPI_num") or 2)
    delai_ev  = float(d.get("Délai_évacuation_min") or 30)
    coord     = int(d.get("Coord_secours") or 2)
    capa_num  = int(d.get("Capacité_médicale_num") or 2)

    # ── Rayons ERPG ────────────────────────────────────────────────────────
    r1 = rayon_seuil(Q_kg, u, stab, H, ERPG["ERPG-1"], dur, type_lib)
    r2 = rayon_seuil(Q_kg, u, stab, H, ERPG["ERPG-2"], dur, type_lib)
    r3 = rayon_seuil(Q_kg, u, stab, H, ERPG["ERPG-3"], dur, type_lib)

    # Zone d'impact et périmètre
    if d.get("Zone_impact_km2") is None:
        d["Zone_impact_km2"] = round(math.pi * (r2 / 1000) ** 2 * (1/6), 3)
    if d.get("Périmètre_évacuation_m") is None:
        d["Périmètre_évacuation_m"] = round(r2, 0)

    # ── Monte Carlo ─────────────────────────────────────────────────────────
    mc = simulation_monte_carlo(Q_kg, u, stab, H, dur, type_lib, dist_pop)
    d["_mc_C_det_ppm"]   = mc["C_det_ppm"]
    d["_mc_niveau_det"]  = mc["niveau_det"]
    d["_mc_p_e1"]        = mc["p_e1"]
    d["_mc_p_e2"]        = mc["p_e2"]
    d["_mc_p_e3"]        = mc["p_e3"]
    d["_mc_niveau_mc"]   = mc["niveau_mc"]
    d["_mc_p50"]         = mc["p50"]
    d["_mc_p95"]         = mc["p95"]
    d["_mc_paradoxe"]    = mc["paradoxe_detecte"]
    d["_rayon_ERPG1_m"]  = round(r1, 0)
    d["_rayon_ERPG2_m"]  = round(r2, 0)
    d["_rayon_ERPG3_m"]  = round(r3, 0)

    # ── Indice de gravité G ─────────────────────────────────────────────────
    G = calcul_indice_G(Q_kg, u, stab, H, dur, type_lib, dist_pop, dens_pop,
                        config, alerte, epi_num, delai_ev, coord, capa_num)
    if d.get("Indice_gravité") is None:
        d["Indice_gravité"] = G
    d["_G_calcule"] = G

    # ── Features ML ────────────────────────────────────────────────────────
    d["_log_Q"]       = round(math.log10(max(Q_kg, 1)), 4)
    d["_log_dist"]    = round(math.log10(max(dist_pop, 1)), 4)
    d["_log_C_proxy"] = round(math.log10(max(
        Q_kg / (u * dist_pop ** 1.5 + 1), 1e-6
    )), 4)
    d["_dens_dist"]   = round(dens_pop / max(dist_pop, 1), 4)

    return d


# ══════════════════════════════════════════════════════════════════════════════
# 6. GESTION DU CACHE ET DE L'HISTORIQUE
# ══════════════════════════════════════════════════════════════════════════════

def load_cache() -> dict:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    return {"seen": [], "last_run": None}


def save_cache(cache: dict):
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def load_history() -> list:
    if LOG_PATH.exists():
        return json.loads(LOG_PATH.read_text(encoding="utf-8"))
    return []


def save_history(history: list):
    LOG_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2, default=str),
                        encoding="utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# 7. ÉCRITURE DANS LA BASE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

# Colonnes officielles de la base SysChlore (ordre identique)
COLS_DB = [
    "Nom_Accident","Année","Pays","Quantité_Cl2_kg","Type_libération",
    "Type_libération_num","Hauteur_source_m","Durée_libération_min",
    "Vitesse_vent_ms","Stabilité_atm","Stabilité_atm_code","Configuration_site",
    "Obstacles","Obstacles_bin","Densité_pop_km2","Distance_pop_m","Durée_expo_min",
    "Niveau_EPI","Niveau_EPI_num","Alerte_précoce","Alerte_précoce_bin",
    "Délai_évacuation_min","Formation_pers","Coord_secours","Délai_intervention_min",
    "Capacité_médicale","Capacité_médicale_num","Décès","Blessés_graves",
    "Blessés_légers","Blessés_total","Zone_impact_km2","Périmètre_évacuation_m",
    "Impact_environnement","Impact_environnement_num","Contamination_eau",
    "Contamination_eau_bin","Contamination_sol","Contamination_sol_bin",
    "Coût_MUSD","Jours_arrêt","Indice_gravité",
]

# Colonnes SysChlore supplémentaires (outputs calculés — feuille séparée)
COLS_SYSCHLORE = [
    "Nom_Accident","Année","Pays",
    "_mc_C_det_ppm","_mc_niveau_det","_mc_p_e1","_mc_p_e2","_mc_p_e3",
    "_mc_niveau_mc","_mc_p50","_mc_p95","_mc_paradoxe",
    "_rayon_ERPG1_m","_rayon_ERPG2_m","_rayon_ERPG3_m","_G_calcule",
    "_log_Q","_log_dist","_log_C_proxy","_dens_dist",
    "source_url","confiance",
]


def ajouter_accident(record: dict, dry_run: bool = False) -> bool:
    """
    Ajoute un accident validé à la base Excel.
    Retourne True si l'ajout a réussi.
    """
    if not DATA_PATH.exists():
        log.error(f"Base de données introuvable : {DATA_PATH}")
        return False
    # Lire la base actuelle
    df_main = pd.read_excel(DATA_PATH, sheet_name="Données", header=1)
    # Vérifier doublon sur Nom + Année
    nom  = record.get("Nom_Accident", "")
    annee = record.get("Année")
    if not df_main.empty:
        mask = (df_main["Nom_Accident"].astype(str) == str(nom)) & \
               (df_main["Année"].astype(str) == str(annee))
        if mask.any():
            log.info(f"  ⏭  Doublon détecté, ignoré : {nom} ({annee})")
            return False
    # Construire la ligne
    row = {c: record.get(c) for c in COLS_DB}
    new_row = pd.DataFrame([row])
    df_updated = pd.concat([df_main, new_row], ignore_index=True)
    # Construire la ligne SysChlore
    row_sc = {c: record.get(c) for c in COLS_SYSCHLORE}
    if dry_run:
        log.info(f"  [DRY-RUN] Ajout simulé : {nom} ({annee})")
        log.info(f"    → G={record.get('_G_calcule')} | MC={record.get('_mc_niveau_mc')} | Det={record.get('_mc_niveau_det')}")
        return True
    # Écrire le fichier Excel (en préservant la feuille Légende)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA_PATH)
        # Feuille Données
        ws_data = wb["Données"]
        # Trouver la dernière ligne remplie
        last_row = ws_data.max_row + 1
        for col_idx, col_name in enumerate(COLS_DB, start=1):
            ws_data.cell(row=last_row, column=col_idx, value=row.get(col_name))
        # Feuille SysChlore_Patch (créer si inexistante)
        if "SysChlore_Patch" not in wb.sheetnames:
            ws_sc = wb.create_sheet("SysChlore_Patch")
            for ci, cn in enumerate(COLS_SYSCHLORE, start=1):
                ws_sc.cell(row=1, column=ci, value=cn)
            sc_start = 2
        else:
            ws_sc = wb["SysChlore_Patch"]
            sc_start = ws_sc.max_row + 1
        for ci, cn in enumerate(COLS_SYSCHLORE, start=1):
            ws_sc.cell(row=sc_start, column=ci, value=row_sc.get(cn))
        wb.save(DATA_PATH)
        log.info(f"  ✅ Ajouté : {nom} ({annee}) — G={record.get('_G_calcule')} | MC={record.get('_mc_niveau_mc')}")
        return True
    except Exception as e:
        log.error(f"Erreur écriture Excel : {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# 8. VALIDATION INTERACTIVE
# ══════════════════════════════════════════════════════════════════════════════

def afficher_fiche(record: dict):
    """Affiche une fiche de synthèse de l'accident extrait."""
    print("\n" + "═" * 68)
    print(f"  ACCIDENT : {record.get('Nom_Accident','?')}  ({record.get('Pays','?')}, {record.get('Année','?')})")
    print("═" * 68)
    print(f"  Source    : {record.get('source_url','?')[:70]}")
    print(f"  Confiance : {record.get('confiance','?')}")
    print()
    print("  ── PARAMÈTRES PHYSIQUES ──────────────────────────────────────")
    print(f"  Quantité Cl₂   : {record.get('Quantité_Cl2_kg')} kg")
    print(f"  Type libération: {record.get('Type_libération')}")
    print(f"  Durée          : {record.get('Durée_libération_min')} min")
    print(f"  Vent           : {record.get('Vitesse_vent_ms')} m/s  |  Stabilité : {record.get('Stabilité_atm')}")
    print(f"  Hauteur source : {record.get('Hauteur_source_m')} m")
    print()
    print("  ── CONTEXTE ──────────────────────────────────────────────────")
    print(f"  Site           : {record.get('Configuration_site')}")
    print(f"  Densité pop    : {record.get('Densité_pop_km2')} hab/km²")
    print(f"  Distance pop   : {record.get('Distance_pop_m')} m")
    print(f"  EPI / Alerte   : {record.get('Niveau_EPI')} / {record.get('Alerte_précoce')}")
    print(f"  Délai évac     : {record.get('Délai_évacuation_min')} min")
    print(f"  Capacité méd   : {record.get('Capacité_médicale')}")
    print()
    print("  ── BILAN HUMAIN ──────────────────────────────────────────────")
    print(f"  Décès          : {record.get('Décès')}  |  Blessés : {record.get('Blessés_total')}")
    print(f"  Coût           : {record.get('Coût_MUSD')} M USD  |  Jours arrêt : {record.get('Jours_arrêt')}")
    print()
    print("  ── RÉSULTATS SysChlore ────────────────────────────────────────")
    mc_par = record.get('_mc_paradoxe', False)
    print(f"  Rayon ERPG-3   : {record.get('_rayon_ERPG3_m')} m")
    print(f"  Rayon ERPG-2   : {record.get('_rayon_ERPG2_m')} m")
    print(f"  C déterministe : {record.get('_mc_C_det_ppm')} ppm  → Alerte det : {record.get('_mc_niveau_det')}")
    print(f"  P(C>ERPG-2) MC : {record.get('_mc_p_e2',0)*100:.1f} %  → Alerte MC  : {record.get('_mc_niveau_mc')}")
    print(f"  P50 / P95      : {record.get('_mc_p50')} / {record.get('_mc_p95')} ppm")
    print(f"  Indice G       : {record.get('_G_calcule')}  (déclaré : {record.get('Indice_gravité')})")
    if mc_par:
        print(f"\n  *** PARADOXE DÉTECTÉ : déterministe={record.get('_mc_niveau_det')} → MC={record.get('_mc_niveau_mc')} ***")
    print("═" * 68)


def valider_interactif(record: dict) -> bool | None:
    """
    Mode interactif : affiche la fiche et demande validation.
    Retourne True (accepter), False (rejeter), None (passer/quitter).
    """
    afficher_fiche(record)
    print("\n  [O] Accepter et ajouter   [N] Rejeter   [M] Modifier   [Q] Quitter")
    while True:
        choix = input("  Votre choix : ").strip().upper()
        if choix == "O":
            return True
        elif choix == "N":
            return False
        elif choix == "Q":
            return None
        elif choix == "M":
            record = modifier_interactif(record)
            afficher_fiche(record)
            print("\n  [O] Accepter   [N] Rejeter   [Q] Quitter")
        else:
            print("  Choix invalide. O / N / M / Q")


def modifier_interactif(record: dict) -> dict:
    """Permet de corriger les champs avant insertion."""
    print("\n  Champs modifiables (Entrée = garder la valeur actuelle) :")
    modifiables = [
        "Quantité_Cl2_kg", "Type_libération", "Durée_libération_min",
        "Vitesse_vent_ms", "Stabilité_atm", "Configuration_site",
        "Densité_pop_km2", "Distance_pop_m", "Décès", "Blessés_total",
        "Niveau_EPI", "Alerte_précoce", "Délai_évacuation_min",
        "Capacité_médicale", "Coût_MUSD", "Indice_gravité",
    ]
    for field in modifiables:
        current = record.get(field)
        val = input(f"  {field} [{current}] : ").strip()
        if val:
            # Convertir en nombre si possible
            try:
                record[field] = float(val) if "." in val else int(val)
            except ValueError:
                record[field] = val
    # Recalculer les outputs après modification
    record = completer_champs_derives(record)
    return record


# ══════════════════════════════════════════════════════════════════════════════
# 9. POINT D'ENTRÉE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def run(auto: bool = False, dry_run: bool = False):
    log.info("══════════════════════════════════════════════")
    log.info(f"SysChlore Weekly Patch — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log.info(f"Mode : {'AUTO' if auto else 'INTERACTIF'}{' DRY-RUN' if dry_run else ''}")
    log.info("══════════════════════════════════════════════")

    # 1. Scraping
    candidates = scrape_all_sources()
    filtered   = filter_chlorine_accidents(candidates)
    if not filtered:
        log.info("✓ Aucun nouvel accident détecté cette semaine.")
        return

    # 2. Extraction et traitement
    cache    = load_cache()
    history  = load_history()
    added    = 0
    rejected = 0
    errors   = 0

    for art in filtered:
        log.info(f"\nTraitement : {art.get('title','?')[:80]}")
        log.info(f"  URL : {art.get('url','?')}")

        # Extraction structurée via Claude
        record = extract_with_claude(art)
        if record is None:
            log.warning("  ✗ Extraction impossible — article ignoré")
            errors += 1
            continue

        # Compléter les champs dérivés et outputs SysChlore
        record = completer_champs_derives(record)

        if auto:
            # Mode automatique : ajouter si confiance Haute ou Moyenne + paradoxe
            confiance = record.get("confiance", "Faible")
            if confiance == "Faible":
                log.info(f"  ⏭  Confiance Faible — ignoré en mode auto")
                rejected += 1
                continue
            ok = ajouter_accident(record, dry_run=dry_run)
        else:
            # Mode interactif
            decision = valider_interactif(record)
            if decision is None:
                log.info("Arrêt demandé par l'utilisateur.")
                break
            elif decision:
                ok = ajouter_accident(record, dry_run=dry_run)
            else:
                log.info(f"  ✗ Rejeté par l'utilisateur : {record.get('Nom_Accident','?')}")
                ok = False
                rejected += 1

        if ok:
            added += 1
            history.append({
                "date_ajout":    datetime.now().isoformat(),
                "Nom_Accident":  record.get("Nom_Accident"),
                "Année":         record.get("Année"),
                "Pays":          record.get("Pays"),
                "source_url":    record.get("source_url"),
                "confiance":     record.get("confiance"),
                "G_calcule":     record.get("_G_calcule"),
                "niveau_mc":     record.get("_mc_niveau_mc"),
                "niveau_det":    record.get("_mc_niveau_det"),
                "paradoxe":      record.get("_mc_paradoxe"),
            })
        # Marquer comme traité dans le cache
        cache["seen"].append(art.get("_hash", ""))

    # 3. Mise à jour cache et historique
    cache["last_run"] = datetime.now().isoformat()
    cache["seen"] = list(set(cache["seen"]))  # dédoublonner
    save_cache(cache)
    save_history(history)

    # 4. Rapport final
    log.info("\n══════════════ RAPPORT FINAL ══════════════")
    log.info(f"Candidats scrappés   : {len(filtered)}")
    log.info(f"Ajoutés à la base    : {added}")
    log.info(f"Rejetés              : {rejected}")
    log.info(f"Erreurs extraction   : {errors}")
    if added > 0:
        log.info(f"Base mise à jour     : {DATA_PATH}")
    log.info("═══════════════════════════════════════════")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SysChlore Weekly Patch")
    parser.add_argument("--auto",    action="store_true",
                        help="Mode automatique sans validation manuelle")
    parser.add_argument("--dry-run", action="store_true",
                        help="Simulation sans écriture dans la base")
    args = parser.parse_args()
    run(auto=args.auto, dry_run=args.dry_run)
